# -*- coding: utf-8 -*-
"""
Created on Thu Nov  5 15:22:39 2020

@author: James Ang
"""

#% Set Directory
import os, sys, subprocess
os.chdir(r'C:\Users\User\Documents\ghub_acceval\smarttradzt-python-services\WS_specialchem')

import re
import time
import openpyxl
import pandas as pd
from datetime import datetime

from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException
# import pkg_resources.py2_warn

import requests
from bs4 import BeautifulSoup

from lib_scrape import append_df_to_excel



#%% Set URL
# login = 'james.ang@acceval-intl.com'
login = 'james.ang@smarttradzt.com'
pw = 'J@mes123'

login_page = 'https://omnexus.specialchem.com/login'

# urls = pd.read_excel('input.xlsx', sheet_name='Sheet1', usecols='A')

#%% Linkedin: Set Driver and open webpage
print('\nMREPC: Open URL...')
chrome_driver = r'C:\Users\User\Documents\ghub_acceval\smarttradzt-python-services\chromedriver.exe'

from selenium.webdriver.chrome.options import Options

options = Options()
options.headless = False

if options.headless == True:
    print ("Headless Chrome Initialized on Linux")
    options.add_argument('--disable-gpu')
else:
    print ("Non-Headless")
    
# options.add_argument("--window-size=1920,1080");
options.add_argument("--start-maximized");

#%% login
driver = webdriver.Chrome(chrome_driver, options=options)
driver.get(login_page)

WebDriverWait(driver,10).until(EC.presence_of_all_elements_located((By.CLASS_NAME,'inbl')))

username = driver.find_element_by_id('txtUsernamelogin')
username.send_keys(login)

password = driver.find_element_by_id('txtPasswordlogin')
password.send_keys(pw)

login_button = driver.find_element_by_id('loginbtnclick')
login_button.click()

WebDriverWait(driver,10).until(EC.presence_of_all_elements_located((By.CLASS_NAME,'inbl')))

#%% Initialisation
product_name = []
product_description = []
product_detail = []
product_properties = []

# product_type = []

# contact_person = []
# e_mail = []
# contact_number = []
# fax = []
# website = []
# company_profile = []
# category_name = []
product_urls = []

count= 0
page = 1

#%% Create new workbook
wb = openpyxl.Workbook()

# Save created workbook at same path where .py file exist
filename = pd.read_excel('input.xlsx', sheet_name='Sheet1', usecols='E',nrows=1).iloc[0,0] +".xlsx"

wb.save(filename)

writer = pd.ExcelWriter(filename, engine = 'xlsxwriter')

#%% 

urls = pd.read_excel('input.xlsx', sheet_name='Sheet1', usecols='A')
url = urls.iloc[0,0]
# url = 'https://omnexus.specialchem.com/selectors/c-thermoplastics-ionomer'
time.sleep(2)
driver.get(url)

WebDriverWait(driver,10).until(EC.presence_of_all_elements_located((By.CLASS_NAME,'box_product')))

#%% : Scraping Products
# for url in urls.iloc[:,0]:
    # print(url)
    # driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)

containers = driver.find_elements_by_class_name('box_product')

# Scrape product URLs
for i in range(0,len(containers)):
    # print(i)
    container = containers[i].find_element_by_class_name('col.w60.zone2.small-w100.medium-w100')
    product_urls.append(container.find_element_by_tag_name('a').get_attribute('href'))


# Scrape individual products

for prod_url in product_urls[:3]:
    
    driver.get(prod_url)
    
    # Product Name
    prod_name = driver.find_element_by_class_name('foro.c9').text
    print('\n')
    print('Scraping ' + prod_name +'...')
    
# =============================================================================
#     # product_name.append(prod_name)
# =============================================================================
    prod_name_df = pd.DataFrame({'Product Name': [prod_name]})
    
# =============================================================================
#     # Product Description
# =============================================================================
    prod_desc = driver.find_element_by_class_name('box_comments.box_limitable.limited').text
    
    # product_description.append(driver.find_element_by_class_name('box_comments.box_limitable.limited').text)
    prod_desc_df = pd.DataFrame({'Product Description': [prod_desc]})
    
    
# =============================================================================
#     # Product Details
# =============================================================================
    product_detail_table = driver.find_element_by_class_name('box_product_details')
    pdt_all_rows = product_detail_table.find_elements_by_class_name('row.tds.pr')
    
    # Scraping pdt rows
    pdt_cat = []
    pdt_val = []
    for row in pdt_all_rows:
        pdt_cat.append(row.find_element_by_class_name('col.w40.small-w100').text)

        pdt_val.append(row.find_element_by_class_name('col.w60.small-w100').text)

    pdt_df = pd.DataFrame({'Product Details Header': pdt_cat, 'Product Details Data':pdt_val})
    
    # Append to excel file
    append_df_to_excel(filename, prod_name_df, sheet_name = f'{prod_name}')
    append_df_to_excel(filename, prod_desc_df, sheet_name = f'{prod_name}')
    append_df_to_excel(filename, pdt_df, sheet_name = f'{prod_name}')
    
# =============================================================================
#     # Product Properties
# =============================================================================
    pp_start = driver.find_element_by_class_name('tds_properties')
    
    if pp_start.text == '':
        
        print('No Product Properties for ' + prod_name + '...')
        
    else:            
    
        pp_title = pp_start.find_element_by_class_name('tds_titre.c9').text
        pp_box = pp_start.find_element_by_class_name('box_content')
        
        pp_tables = pp_box.find_elements_by_class_name('box_limitable.limited')
        
        ppt_df = {}
        i = 0
        
        # Scraping Individual Property tables
        for table in pp_tables:
            # print('\n\t' + table.text)
            pp_t1 = table.find_elements_by_class_name('tr.h100')
            pp_t1_allheaders = pp_t1[0].find_elements_by_class_name('row.w100.h100')
            pp_t1_heading1 = pp_t1_allheaders[0].find_element_by_class_name('col.w70.medium-w66.fs18.small-w100.pt15').text
            pp_t1_heading2 = pp_t1_allheaders[0].find_element_by_class_name('col.w30.medium-w33.c13.small-hidden.pt15').text
            pp_t1_heading3 = pp_t1_allheaders[1].find_element_by_class_name('col.w60.c13.medium-hidden.small-hidden.pt15').text
            pp_t1_heading4 =pp_t1_allheaders[1].find_element_by_class_name('col.w30.c13.medium-hidden.small-hidden.pt15').text
            
            # Scrape Entries in pp table 1
            pp_t1_en = table.find_element_by_class_name('container_properties')    
            pp_t1_entries = pp_t1_en.find_elements_by_class_name('row.tds.pr')
            
            h1_data = []
            h2_data = []
            h3_data = []
            h4_data = []
            
            for item in pp_t1_entries:
                # print('\n\t\t' + item.text)
                h1_data.append(item.find_element_by_class_name('col.w70.medium-w66.small-w100').text)
                h2_data.append(item.find_element_by_class_name('col.w30.medium-w33.small-w100.value-unit').text)
                h3_data.append(item.find_element_by_class_name('col.large-w70.medium-inbl.small-inbl.test-condition').text)
                h4_data.append(item.find_element_by_class_name('col.large-w30.medium-inbl.small-inbl.test-method').text)
                
            ppt_df[i] = pd.DataFrame({
                pp_t1_heading1: h1_data,
                pp_t1_heading2: h2_data,
                pp_t1_heading3: h3_data,
                pp_t1_heading4: h4_data,
                })
            ppt_df[i].head()
            
            # Append to excel file
            append_df_to_excel(filename, ppt_df[i], sheet_name=f'{prod_name}')
            
            
            
            
            
            
            
            i +=1

# driver.close()
        #%%
#         pp_t2 = pp_tables[1].find_elements_by_class_name('tr.h100') # Variable
#         pp_t2_allheaders = pp_t2[0].find_elements_by_class_name('row.w100.h100')
#         pp_t2_heading1 = pp_t2_allheaders[0].find_element_by_class_name('col.w70.medium-w66.fs18.small-w100.pt15').text
#         pp_t2_heading2 = pp_t2_allheaders[0].find_element_by_class_name('col.w30.medium-w33.c13.small-hidden.pt15').text
#         pp_t2_heading3 = pp_t2_allheaders[1].find_element_by_class_name('col.w60.c13.medium-hidden.small-hidden.pt15').text
#         pp_t2_heading4 =pp_t2_allheaders[1].find_element_by_class_name('col.w30.c13.medium-hidden.small-hidden.pt15').text
        
#         # Scrape Entries in pp table 1
#         pp_t2_en = pp_tables[1].find_element_by_class_name('container_properties')    # Variable
#         pp_t2_entries = pp_t2_en.find_elements_by_class_name('row.tds.pr')
        
#         h1_data = []
#         h2_data = []
#         h3_data = []
#         h4_data = []
        
#         for item in pp_t2_entries:
#             print(item.text)
#             h1_data.append(item.find_element_by_class_name('col.w70.medium-w66.small-w100').text)
#             h2_data.append(item.find_element_by_class_name('col.w30.medium-w33.small-w100.value-unit').text)
#             h3_data.append(item.find_element_by_class_name('col.large-w70.medium-inbl.small-inbl.test-condition').text)
#             h4_data.append(item.find_element_by_class_name('col.large-w30.medium-inbl.small-inbl.test-method').text)
            
#         ppt2_df = pd.DataFrame({
#             pp_t2_heading1: h1_data,
#             pp_t2_heading2: h2_data,
#             pp_t2_heading3: h3_data,
#             pp_t2_heading4: h4_data,
#             })
        
#         # Save to Excel
#         append_df_to_excel(filename, ppt1_df, sheet_name=f'{prod_name}')
    


# #%%
# page = requests.get(product_urls[3]),

# driver.get(product_urls[3])

# page = requests.get("https://omnexus.specialchem.com/product/t-braskem-fl100pp")
# page = requests.get("https://omnexus.specialchem.com/product/t-arkema-n3xtdimension-n3d-f2110")
# page = requests.get("https://omnexus.specialchem.com/product/t-dsm-arnite-id-3040")

# soup = BeautifulSoup(page.content, 'html.parser')
# print(soup.prettify())

# # temp =list(soup.children)
# # temp = [type(item) for item in list(soup.children)]
# # html = temp[3]
# # list(html.children)
# # soup.find_all('p')

# # Product Name
# product_name.append(soup.find('h1', class_='foro c9').text)

# # Product Description
# prod_desc = soup.find('div', class_='box_comments box_limitable limited').text
# product_description.append(prod_desc.split('\n')[1])

# # Box product details
# soup.find("div", attrs={"class": "box_product_details"})
# table_top = soup.find_all('div', class_='box_product_details')
# table_top.div.find_all("div", attrs={"class": "row tds pr"})
# table_top.find_all('div', class_='row tds pr')
# table_top.find_all('div', class_='tr')




# # Product type
# table_top = soup.find_all('div', class_='col w60 small-w100')
# prod_type = table_top[0].text.split('\n')
# product_type.append(",".join(string for string in prod_type if len(string) > 0))


# # table bottom
# table_bot = soup.find('div', class_='tds_properties')
# table_bot.find('div',class_='tds_titre c9')
# table_bot.find('div',class_='box_content')
 




# # e = d[0].find_all('li')
# # product_type.append(e[0].text.rstrip("\n").lstrip("\n") + '\n'+ e[1].text.rstrip("\n").lstrip("\n"))





# name = driver.find_element_by_class_name('crumb.global')
# cat_name = name.text.split(' > ')[-1].split(' | ')[0]

# if 'Search' in cat_name:
#     cat_name = cat_name.split(" : ")[-1]
    
# print('\n' + cat_name)
# f = driver.find_elements_by_class_name('box4')
# # g = f.find_elements_by_class_name("detail")

# page = driver.find_elements_by_class_name('page_count')

# if  page ==[] or page[0].text =='':
#     print('1 page')
#     num_page = 0               
    
#     for item in f:
#         count += 1
#         print(f'Item {count}')
        
#         category_name.append(cat_name)
#         companies.append(item.text.split('\n')[0])
        
#         # Contact person
#         if len(item.text.split('\n')[1].split(': ')) > 1:
#             contact_person.append(item.text.split('\n')[1].split(': ')[1])
#         else:
#             contact_person.append('NA')
        
#         # E-mail
#         if len(item.text.split('\n')[2].split(': ')) > 1:
#             e_mail.append(item.text.split('\n')[2].split(': ')[1])
#         else:
#             e_mail.append('NA')
            
#         # Contact Number
#         if len(item.text.split('\n')[3].split(': ')) > 1:
#             contact_number.append(item.text.split('\n')[3].split(': ')[1])
#         else:
#             contact_number.append('NA')
        
#         # fax
#         if len(item.text.split('\n')[4].split(': ')) > 1:
#             fax.append(item.text.split('\n')[4].split(': ')[1])
#         else:
#             fax.append('NA')
                    
#         # website
#         if len(item.text.split('\n')[5].split(': ')) > 1:
#             website.append(item.text.split('\n')[5].split(': ')[1])
#         else:
#             website.append('NA')
    
# else:
#     print('More than 2 pages')        
    
#     for item in f:
#         count += 1
#         print(f'Item {count}')
        
#         category_name.append(cat_name)
#         companies.append(item.text.split('\n')[0])
        
#         # Contact person
#         if len(item.text.split('\n')[1].split(': ')) > 1:
#             contact_person.append(item.text.split('\n')[1].split(': ')[1])
#         else:
#             contact_person.append('NA')
        
#         # E-mail
#         if len(item.text.split('\n')[2].split(': ')) > 1:
#             e_mail.append(item.text.split('\n')[2].split(': ')[1])
#         else:
#             e_mail.append('NA')
            
#         # Contact Number
#         if len(item.text.split('\n')[3].split(': ')) > 1:
#             contact_number.append(item.text.split('\n')[3].split(': ')[1])
#         else:
#             contact_number.append('NA')
        
#         # fax
#         if len(item.text.split('\n')[4].split(': ')) > 1:
#             fax.append(item.text.split('\n')[4].split(': ')[1])
#         else:
#             fax.append('NA')
                    
#         # website
#         if len(item.text.split('\n')[5].split(': ')) > 1:
#             website.append(item.text.split('\n')[5].split(': ')[1])
#         else:
#             website.append('NA')
            
#         # contact_person.append(item.text.split('\n')[1].split(': ')[1])
#         # e_mail.append(item.text.split('\n')[2].split(': ')[1])
#         # contact_number.append(item.text.split('\n')[3].split(': ')[1])
#         # fax.append(item.text.split('\n')[4].split(': ')[1])
    
#     num_page = int(page[-2].text)
    
#     # Pagination
#     for i in range(0,num_page-1):
#         print(i)
#         url_new = page[-1].get_attribute('href')
#         driver.get(url_new)
#         name = driver.find_element_by_class_name('crumb.global')
#         cat_name = name.text.split(' > ')[-1].split(' | ')[0]
        
#         if 'Search' in cat_name:
#             cat_name = cat_name.split(" : ")[-1]
        
#         f = driver.find_elements_by_class_name('box4')
#         page = driver.find_elements_by_class_name('page_count')

#         # Scraping items
#         for item in f:
#             count += 1
#             print(f'Item {count}')
            
#             category_name.append(cat_name)
#             companies.append(item.text.split('\n')[0])
            
#             # Contact person
#             if len(item.text.split('\n')[1].split(': ')) > 1:
#                 contact_person.append(item.text.split('\n')[1].split(': ')[1])
#             else:
#                 contact_person.append('NA')
            
#             # E-mail
#             if len(item.text.split('\n')[2].split(': ')) > 1:
#                 e_mail.append(item.text.split('\n')[2].split(': ')[1])
#             else:
#                 e_mail.append('NA')
                
#             # Contact Number
#             if len(item.text.split('\n')[3].split(': ')) > 1:
#                 contact_number.append(item.text.split('\n')[3].split(': ')[1])
#             else:
#                 contact_number.append('NA')
            
#             # fax
#             if len(item.text.split('\n')[4].split(': ')) > 1:
#                 fax.append(item.text.split('\n')[4].split(': ')[1])
#             else:
#                 fax.append('NA')
            
#             # website
#             if len(item.text.split('\n')[5].split(': ')) > 1:
#                 website.append(item.text.split('\n')[5].split(': ')[1])
#             else:
#                 website.append('NA')
                
    
#     # for item in f:
#     #     count += 1
#     #     print(f'Item {count}')
        
#     #     category_name.append(cat_name)
#     #     companies.append(item.text.split('\n')[0])
#     #     contact_person.append(item.text.split('\n')[1].split(': ')[1])
#     #     e_mail.append(item.text.split('\n')[2].split(': ')[1])
#     #     contact_number.append(item.text.split('\n')[3].split(': ')[1])
#     #     fax.append(item.text.split('\n')[4].split(': ')[1])
#     #     website.append(item.text.split('\n')[5].split(': ')[1])
    
#      #%% Close website
# driver.close()

# #%% Compile and Prepare Data for export
# df = pd.DataFrame({'Product Category': category_name, 'Company':companies, 'Contact Person': contact_person, 'E-mail': e_mail, 'Contact Number':contact_number, 'Fax':fax, 'Website':website})

# import openpyxl
# # Create new workbook
# wb = openpyxl.Workbook()

# # Get SHEET name
# # Sheet_name = wb.sheetnames

# # Save created workbook at same path where .py file exist
# filename = pd.read_excel('input.xlsx', sheet_name='Sheet1', usecols='E',nrows=1).iloc[0,0] +".xlsx"

# wb.save(filename)
# # dateTimeObj = datetime.now()
# # timestampStr = dateTimeObj.strftime("%d%b%Y_%H%M%S")

# # df.to_excel(filename+'.xlsx', index=False, encoding='utf-8')

# # filename = r'C:/Users/User/Documents/ghub_acceval/smarttradzt-python-services/WS_mrepc/MREPC_.xlsx'
# append_df_to_excel(filename, df)

